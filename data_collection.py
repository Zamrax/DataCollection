import openpyxl


# Increase the data in the sheet by 1
def increase_data(sheet, dsa):
    if dsa == "Yes" or dsa == "yes" or dsa == "Y" or dsa == "y":
        print('What type of cigarette? (Cigarettes, Vape, Juul, Iqos, Other)')
        type = input()
        if type == "Cigarettes":
            sheet['B2'].value += 1
        elif type == "Vape":
            sheet['B3'].value += 1
        elif type == "Juul":
            sheet['B4'].value += 1
        elif type == "Iqos":
            sheet['B5'].value += 1
        elif type == "Other":
            sheet['B6'].value += 1
        else:
            return 1
        return 0
    elif dsa == "No" or dsa == "no" or dsa == "N" or dsa == "n":
        print('What type of cigarette? (Cigarettes, Vape, Juul, Iqos, Other)')
        type = input()
        if type == "Cigarettes":
            sheet['C2'].value += 1
        elif type == "Vape":
            sheet['C3'].value += 1
        elif type == "Juul":
            sheet['C4'].value += 1
        elif type == "Iqos":
            sheet['C5'].value += 1
        elif type == "Other":
            sheet['C6'].value += 1
        else:
            return 1
        return 0
    else:
        return 1


# Decrease the data in the sheet by 1
def decrease_data(sheet, dsa):
    if dsa == "Yes" or dsa == "yes" or dsa == "Y" or dsa == "y":
        print('What type of cigarette? (Cigarettes, Vape, Juul, Iqos, Other)')
        type = input()
        if type == "Cigarettes":
            if sheet['B2'].value == 0:
                print("Cannot go below 0")
                return 0
            sheet['B2'].value -= 1
        elif type == "Vape":
            if sheet['B3'].value == 0:
                print("Cannot go below 0")
                return 0
            sheet['B3'].value -= 1
        elif type == "Juul":
            if sheet['B4'].value == 0:
                print("Cannot go below 0")
                return 0
            sheet['B4'].value -= 1
        elif type == "Iqos":
            if sheet['B5'].value == 0:
                print("Cannot go below 0")
                return 0
            sheet['B5'].value -= 1
        elif type == "Other":
            if sheet['B6'].value == 0:
                print("Cannot go below 0")
                return 0
            sheet['B6'].value -= 1
        else:
            return 1
        return 0
    elif dsa == "No" or dsa == "no" or dsa == "N" or dsa == "n":
        print('What type of cigarette? (Cigarettes, Vape, Juul, Iqos, Other)')
        type = input()
        if type == "Cigarettes":
            if sheet['C2'].value == 0:
                print("Cannot go below 0")
                return 0
            sheet['C2'].value -= 1
        elif type == "Vape":
            if sheet['C3'].value == 0:
                print("Cannot go below 0")
                return 0
            sheet['C3'].value -= 1
        elif type == "Juul":
            if sheet['C4'].value == 0:
                print("Cannot go below 0")
                return 0
            sheet['C4'].value -= 1
        elif type == "Iqos":
            if sheet['C5'].value == 0:
                print("Cannot go below 0")
                return 0
            sheet['C5'].value -= 1
        elif type == "Other":
            if sheet['C6'].value == 0:
                print("Cannot go below 0")
                return 0
            sheet['C6'].value -= 1
        else:
            return 1
        return 0
    else:
        return 1


# Add a certain amount to the data in the sheet
def add_data(sheet, dsa, amount):
    if dsa == "Yes" or dsa == "yes" or dsa == "Y" or dsa == "y":
        print('What type of cigarette? (Cigarettes, Vape, Juul, Iqos, Other)')
        type = input()
        if type == "Cigarettes":
            sheet['B2'].value += amount
        elif type == "Vape":
            sheet['B3'].value += amount
        elif type == "Juul":
            sheet['B4'].value += amount
        elif type == "Iqos":
            sheet['B5'].value += amount
        elif type == "Other":
            sheet['B6'].value += amount
        else:
            return 1
        return 0
    elif dsa == "No" or dsa == "no" or dsa == "N" or dsa == "n":
        print('What type of cigarette? (Cigarettes, Vape, Juul, Iqos, Other)')
        type = input()
        if type == "Cigarettes":
            sheet['C2'].value += amount
        elif type == "Vape":
            sheet['C3'].value += amount
        elif type == "Juul":
            sheet['C4'].value += amount
        elif type == "Iqos":
            sheet['C5'].value += amount
        elif type == "Other":
            sheet['C6'].value += amount
        else:
            return 1
        return 0
    else:
        return 1


# Decrease the data in the sheet by certain amount
def delete_data(sheet, dsa, amount):
    if dsa == "Yes" or dsa == "yes" or dsa == "Y" or dsa == "y":
        print('What type of cigarette? (Cigarettes, Vape, Juul, Iqos, Other)')
        type = input()
        if type == "Cigarettes":
            temp = sheet['B2'].value
            if temp - amount < 0:
                print("Cannot go below 0")
                return 0
            sheet['B2'].value -= amount
        elif type == "Vape":
            temp = sheet['B3'].value
            if temp - amount < 0:
                print("Cannot go below 0")
                return 0
            sheet['B3'].value -= amount
        elif type == "Juul":
            temp = sheet['B4'].value
            if temp - amount < 0:
                print("Cannot go below 0")
                return 0
            sheet['B4'].value -= amount
        elif type == "Iqos":
            temp = sheet['B5'].value
            if temp - amount < 0:
                print("Cannot go below 0")
                return 0
            sheet['B5'].value -= amount
        elif type == "Other":
            temp = sheet['B6'].value
            if temp - amount < 0:
                print("Cannot go below 0")
                return 0
            sheet['B6'].value -= amount
        else:
            return 1
        return 0
    elif dsa == "No" or dsa == "no" or dsa == "N" or dsa == "n":
        print('What type of cigarette? (Cigarettes, Vape, Juul, Iqos, Other)')
        type = input()
        if type == "Cigarettes":
            temp = sheet['C2'].value
            if temp - amount < 0:
                print("Cannot go below 0")
                return 0
            sheet['C2'].value -= amount
        elif type == "Vape":
            temp = sheet['C3'].value
            if temp - amount < 0:
                print("Cannot go below 0")
                return 0
            sheet['C3'].value -= amount
        elif type == "Juul":
            temp = sheet['C4'].value
            if temp - amount < 0:
                print("Cannot go below 0")
                return 0
            sheet['C4'].value -= amount
        elif type == "Iqos":
            temp = sheet['C5'].value
            if temp - amount < 0:
                print("Cannot go below 0")
                return 0
            sheet['C5'].value -= amount
        elif type == "Other":
            temp = sheet['C6'].value
            if temp - amount < 0:
                print("Cannot go below 0")
                return 0
            sheet['C6'].value -= amount
        else:
            return 1
        return 0
    else:
        return 1


if __name__ == "__main__":
    print("What file do you want to work with?:")
    file = input()
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    while True:
        print('What do you want to do?')
        command = input()
        if command == "Increase the data" or command == "increase the data" or command == "Increase" or command == "increase":
            print('Designated smoking area?')
            dsa = input()
            result = increase_data(sheet, dsa)
            if result == 1:
                print('Try again')
        elif command == "Decrease the data" or command == "decrease the data" or command == "Decrease" or command == "decrease":
            print('Designated smoking area?')
            dsa = input()
            result = decrease_data(sheet, dsa)
            if result == 1:
                print('Try again')
        elif command == "Add to the data" or command == "add to the data" or command == "Add" or command == "add":
            print('Designated smoking area?')
            dsa = input()
            print('How many units?')
            amount = input()
            result = add_data(sheet, dsa, amount)
            if result == 1:
                print('Try again')
        elif command == "Delete from the data" or command == "delete from the data" or command == "Delete" or command == "delete":
            print('Designated smoking area?')
            dsa = input()
            print('How many units?')
            amount = input()
            result = delete_data(sheet, dsa, amount)
            if result == 1:
                print('Try again')
        elif command == "Print data" or command == "print" or command == "Print":
            for i in range(1, sheet.max_row + 1):
                for j in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row = i, column = j)
                    print(cell.value,end=' | ')
                print('\n')
        elif command == "Help" or command == "help":
            print("Usage:")
            print("Increase data by one:")
            print("Increase the data / increase the data / Increase/ increase\n")
            print("Dcrease data by one:")
            print("Decrease the data / decrease the data / Decrease / decrease\n")
            print("Increase data by certain amount:")
            print("Add to the data / add to the data / Add / add\n")
            print("Decrease data by certain amount:")
            print("Delete from the data / delete from the data / Delete / delete\n")
        elif command == "Exit" or command == "exit":
            break
        else:
            print('Try again!')